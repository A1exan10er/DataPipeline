"""Shared data structures and utilities for the QA pipeline."""

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
import csv
import json
import sqlite3

from scripts.pipeline.qa_config import config_value


DEFAULT_DB_PATH = Path("outputs/qa_pipeline.db")
STATUS_VALUES = ("pass", "warning", "fail", "needs_review")
SEVERITY_VALUES = ("info", "minor", "major", "critical")


class PipelineConfigurationError(RuntimeError):
    """Raised when the pipeline environment is not configured to run safely."""


@dataclass
class Finding:
    """One specific problem or result detected during a check."""

    episode_path: str
    phase: int
    check_name: str
    severity: str
    status: str
    message: str
    details: dict = field(default_factory=dict)


@dataclass
class EpisodeState:
    """Accumulated QA state for one episode across pipeline phases."""

    episode_path: Path
    task: str
    date: str
    operator: str
    robot: str
    controller: str
    metadata: dict = field(default_factory=dict)
    phases_completed: list[int] = field(default_factory=list)
    phase_status: dict = field(default_factory=dict)
    findings: list[Finding] = field(default_factory=list)
    metrics: dict = field(default_factory=dict)
    final_status: str = ""
    training_ready: bool | None = None
    last_updated: str = field(default_factory=lambda: datetime.now().isoformat())


def init_db(db_path: Path) -> None:
    """Create the SQLite database and required tables if they do not exist."""
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS episodes (
                episode_path      TEXT PRIMARY KEY,
                task              TEXT,
                date              TEXT,
                operator          TEXT,
                robot             TEXT,
                controller        TEXT,
                phases_completed  TEXT,
                phase_status      TEXT,
                metrics           TEXT,
                final_status      TEXT,
                training_ready    INTEGER,
                last_updated      TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS findings (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                episode_path  TEXT,
                phase         INTEGER,
                check_name    TEXT,
                severity      TEXT,
                status        TEXT,
                message       TEXT,
                details       TEXT
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_findings_episode_path "
            "ON findings(episode_path)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_findings_phase "
            "ON findings(phase)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_findings_status "
            "ON findings(status)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_episodes_final_status "
            "ON episodes(final_status)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_episodes_task "
            "ON episodes(task)"
        )


def load_episode_state(db_path: Path, episode_path: Path) -> EpisodeState | None:
    """Load one episode state from SQLite, or return None if absent."""
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            """
            SELECT episode_path, task, date, operator, robot, controller,
                   phases_completed, phase_status, metrics, final_status,
                   training_ready, last_updated
            FROM episodes
            WHERE episode_path = ?
            """,
            (str(episode_path),),
        ).fetchone()
    if row is None:
        return None
    return _state_from_row(row)


def save_episode_state(db_path: Path, state: EpisodeState) -> None:
    """Upsert one episode state into SQLite without writing findings."""
    init_db(db_path)
    training_ready = None if state.training_ready is None else int(state.training_ready)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO episodes (
                episode_path, task, date, operator, robot, controller,
                phases_completed, phase_status, metrics, final_status,
                training_ready, last_updated
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(state.episode_path),
                state.task,
                state.date,
                state.operator,
                state.robot,
                state.controller,
                json.dumps(state.phases_completed, ensure_ascii=False),
                json.dumps(state.phase_status, ensure_ascii=False),
                json.dumps(state.metrics, ensure_ascii=False),
                state.final_status,
                training_ready,
                state.last_updated,
            ),
        )


def save_findings(
    db_path: Path,
    findings: list[Finding],
    phase: int | None = None,
    episode_path: str | None = None,
) -> None:
    """Save findings for an episode and phase, replacing any existing findings.

    Always deletes existing findings for the same episode_path and phase before
    inserting new ones, even when new_findings is empty. This ensures reruns
    produce clean results.

    Args:
        db_path: Path to the SQLite database.
        findings: List of Finding objects to save. May be empty.
        phase: Phase number to clear before inserting. If None, infer from
               findings. If findings is empty and phase is None, nothing is
               cleared (caller must supply phase explicitly for empty lists).
        episode_path: Episode path to clear when findings is empty.
    """
    if not findings and phase is None:
        return
    init_db(db_path)
    phase_to_clear = phase if phase is not None else findings[0].phase
    episode_path_to_clear = episode_path or (findings[0].episode_path if findings else None)
    if episode_path_to_clear is None:
        return
    insert_rows = [_finding_insert_row(finding) for finding in findings]
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            DELETE FROM findings
            WHERE episode_path = ? AND phase = ?
            """,
            (episode_path_to_clear, phase_to_clear),
        )
        if insert_rows:
            conn.executemany(
                """
                INSERT INTO findings (
                    episode_path, phase, check_name, severity, status, message, details
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                insert_rows,
            )


def load_all_states(db_path: Path) -> list[EpisodeState]:
    """Load all episode states without loading findings."""
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT episode_path, task, date, operator, robot, controller,
                   phases_completed, phase_status, metrics, final_status,
                   training_ready, last_updated
            FROM episodes
            ORDER BY episode_path
            """,
            (),
        ).fetchall()
    return [_state_from_row(row) for row in rows]


def discover_episodes(roots: list[Path]) -> list[Path]:
    """Discover episode directories below roots, skipping _quarantine trees."""
    episodes: set[str] = set()
    for root in roots:
        root_path = Path(root)
        if root_path.is_dir():
            _discover_under_root(root_path, episodes)
    return [Path(path) for path in sorted(episodes)]


def load_metadata(episode_path: Path) -> tuple[dict, list[Finding]]:
    """Load metadata.json from an episode and return metadata plus findings."""
    metadata_path = Path(episode_path) / "metadata.json"
    if not metadata_path.exists():
        return {}, [_metadata_finding(episode_path, "metadata_exists", "metadata.json is missing")]
    try:
        with metadata_path.open("r", encoding="utf-8") as file_obj:
            data = json.load(file_obj)
    except json.JSONDecodeError as exc:
        return {}, [
            _metadata_finding(
                episode_path,
                "metadata_valid_json",
                "metadata.json is not valid JSON",
                {"error": str(exc)},
            )
        ]
    except OSError as exc:
        return {}, [
            _metadata_finding(
                episode_path,
                "metadata_valid_json",
                "metadata.json could not be read",
                {"error": str(exc)},
            )
        ]
    if not isinstance(data, dict):
        return {}, [
            _metadata_finding(
                episode_path,
                "metadata_valid_json",
                "metadata.json must contain a JSON object",
                {"actual_type": type(data).__name__},
            )
        ]
    return data, []


def infer_context(roots: list[Path], episode_path: Path, metadata: dict) -> dict:
    """Infer task, date, operator, robot, and controller from path and metadata."""
    relative_parts, root = _relative_episode_parts(roots, episode_path)
    date_index = _find_date_index(relative_parts)
    date = relative_parts[date_index] if date_index is not None else ""
    operator = _operator_from_parts(relative_parts, date_index)
    task_folder = _task_from_parts(relative_parts, root, date_index)
    robot = first_present(
        metadata.get("robot"),
        _robot_from_episode_name(Path(episode_path).name),
        _robot_from_layout_parts(relative_parts, date_index),
    )
    controller = first_present(metadata.get("controller"), metadata.get("controller_type"))
    if not controller:
        controller = _controller_from_episode_name(Path(episode_path).name)
    return {
        "task": first_present(metadata.get("task_key"), task_folder),
        "date": date,
        "operator": operator,
        "robot": robot,
        "controller": controller,
    }


def severity_rank(severity: str) -> int:
    """Return a numeric rank for a severity string."""
    return {"critical": 3, "major": 2, "minor": 1, "info": 0}.get(severity, 0)


def max_severity(findings: list[Finding], minimum: str = "info") -> str:
    """Return the highest severity in findings, or minimum for no findings."""
    highest = minimum
    for finding in findings:
        if severity_rank(finding.severity) > severity_rank(highest):
            highest = finding.severity
    return highest


def decide_status(findings: list[Finding]) -> str:
    """Decide an aggregate status from a list of findings."""
    if any(finding.severity == "critical" for finding in findings):
        return "fail"
    if any(finding.severity == "major" and finding.status == "fail" for finding in findings):
        return "fail"
    if any(finding.status == "needs_review" for finding in findings):
        return "needs_review"
    if any(finding.severity == "major" for finding in findings):
        return "warning"
    if any(finding.severity == "minor" for finding in findings):
        return "warning"
    return "pass"


def export_quality_report(db_path: Path, output_path: Path) -> None:
    """Export one CSV report row per episode."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _quality_report_rows(db_path)
    fieldnames = [
        "episode_path",
        "task",
        "date",
        "operator",
        "robot",
        "controller",
        "status",
        "severity",
        "reasons",
        "checked_at",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as file_obj:
        writer = csv.DictWriter(file_obj, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def export_excel_report(db_path: Path, output_path: Path) -> None:
    """Export a human-readable Excel workbook for sharing QA results."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise PipelineConfigurationError(
            "Excel export requires openpyxl. Install it in datapipeline-env with: "
            "python3 -m pip install openpyxl"
        ) from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    episode_rows = _quality_report_rows(db_path)
    finding_rows = _excel_finding_rows(db_path)
    summary_rows = _excel_summary_rows(episode_rows, finding_rows)
    task_rows = _excel_task_rows(episode_rows)
    issue_rows = _excel_issue_count_rows(finding_rows)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_sheet(summary_sheet, summary_rows, ["metric", "value"])
    _write_sheet(workbook.create_sheet("Episodes"), episode_rows, [
        "episode_path", "task", "date", "operator", "robot", "controller",
        "status", "severity", "reasons", "checked_at",
    ])
    _write_sheet(workbook.create_sheet("Findings"), finding_rows, [
        "episode_path", "task", "date", "operator", "robot", "controller",
        "phase", "check_name", "severity", "status", "message", "details",
    ])
    _write_sheet(workbook.create_sheet("Issue Counts"), issue_rows, ["check_name", "count"])
    _write_sheet(workbook.create_sheet("Task Status"), task_rows, [
        "task", "fail", "needs_review", "warning", "pass", "pending", "total",
    ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _autosize_columns(worksheet, get_column_letter)

    workbook.save(output_path)


def export_findings_jsonl(db_path: Path, output_path: Path) -> None:
    """Export all findings as newline-delimited JSON."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    init_db(db_path)
    with sqlite3.connect(db_path) as conn, output_path.open("w", encoding="utf-8") as file_obj:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT episode_path, phase, check_name, severity, status, message, details
            FROM findings
            ORDER BY episode_path, phase, id
            """,
            (),
        )
        for row in rows:
            file_obj.write(json.dumps(_finding_json(row), ensure_ascii=False) + "\n")


def export_summary_md(db_path: Path, output_path: Path) -> None:
    """Export a Markdown summary of quality status and issue patterns."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    init_db(db_path)
    lines = ["# Quality Summary", ""]
    lines.extend(_summary_overview_lines(db_path))
    lines.extend(_summary_by_task_lines(db_path))
    lines.extend(_summary_by_operator_lines(db_path))
    lines.extend(_summary_by_robot_lines(db_path))
    lines.extend(_summary_top_issues_lines(db_path))
    lines.extend(_summary_problem_patterns_lines(db_path))
    lines.extend(_summary_appendix_lines(db_path))
    output_path.write_text("\n".join(lines), encoding="utf-8")


def first_present(*values: Any) -> Any:
    """Return the first value that is not None and not an empty string."""
    for value in values:
        if value is not None and value != "":
            return value
    return ""


def is_positive_number(value: Any) -> bool:
    """Return True when value can be converted to a positive float."""
    try:
        return float(value) > 0
    except (TypeError, ValueError):
        return False


def count_csv_rows(path: Path) -> int | None:
    """Count CSV data rows without loading the full file into memory."""
    try:
        with Path(path).open("r", encoding="utf-8", newline="") as file_obj:
            reader = csv.reader(file_obj)
            next(reader, None)
            return sum(1 for _ in reader)
    except (OSError, csv.Error, UnicodeDecodeError):
        return None


def _state_from_row(row: sqlite3.Row) -> EpisodeState:
    episode_path = Path(row["episode_path"])
    metadata, _ = load_metadata(episode_path)
    return EpisodeState(
        episode_path=episode_path,
        task=row["task"] or "",
        date=row["date"] or "",
        operator=row["operator"] or "",
        robot=row["robot"] or "",
        controller=row["controller"] or "",
        metadata=metadata,
        phases_completed=_json_list(row["phases_completed"]),
        phase_status=_json_dict(row["phase_status"], convert_int_keys=True),
        findings=[],
        metrics=_json_dict(row["metrics"]),
        final_status=row["final_status"] or "",
        training_ready=_db_bool(row["training_ready"]),
        last_updated=row["last_updated"] or "",
    )


def _json_list(value: str | None) -> list[int]:
    try:
        data = json.loads(value or "[]")
    except json.JSONDecodeError:
        return []
    return data if isinstance(data, list) else []


def _json_dict(value: str | None, convert_int_keys: bool = False) -> dict:
    try:
        data = json.loads(value or "{}")
    except json.JSONDecodeError:
        return {}
    if not isinstance(data, dict):
        return {}
    if convert_int_keys:
        return {_maybe_int(key): item for key, item in data.items()}
    return data


def _maybe_int(value: Any) -> Any:
    if isinstance(value, str) and value.isdigit():
        return int(value)
    return value


def _db_bool(value: Any) -> bool | None:
    if value is None:
        return None
    return bool(value)


def _finding_insert_row(finding: Finding) -> tuple:
    return (
        finding.episode_path,
        finding.phase,
        finding.check_name,
        finding.severity,
        finding.status,
        finding.message,
        json.dumps(finding.details, ensure_ascii=False),
    )


def _discover_under_root(root: Path, episodes: set[str]) -> None:
    stack = [root]
    while stack:
        current = stack.pop()
        if current.name == "_quarantine":
            continue
        if current.name.startswith("episode_"):
            episodes.add(str(current))
            continue
        try:
            children = [child for child in current.iterdir() if child.is_dir()]
        except OSError:
            continue
        stack.extend(child for child in children if child.name != "_quarantine")


def _metadata_finding(
    episode_path: Path, check_name: str, message: str, details: dict | None = None
) -> Finding:
    return Finding(
        episode_path=str(episode_path),
        phase=1,
        check_name=check_name,
        severity="critical",
        status="fail",
        message=message,
        details=details or {},
    )


def _relative_episode_parts(roots: list[Path], episode_path: Path) -> tuple[tuple[str, ...], Path | None]:
    episode_resolved = Path(episode_path).resolve(strict=False)
    matches: list[tuple[int, Path, tuple[str, ...]]] = []
    for root in roots:
        root_resolved = Path(root).resolve(strict=False)
        try:
            relative = episode_resolved.relative_to(root_resolved)
        except ValueError:
            continue
        matches.append((len(root_resolved.parts), Path(root), relative.parts))
    if not matches:
        return Path(episode_path).parts, None
    _, root, parts = max(matches, key=lambda item: item[0])
    return parts, root


def _find_date_index(parts: tuple[str, ...]) -> int | None:
    for index, part in enumerate(parts):
        if len(part) == 8 and part.isdigit():
            return index
    return None


def _operator_from_parts(parts: tuple[str, ...], date_index: int | None) -> str:
    if date_index is None:
        return ""
    operator_index = date_index + 1
    return parts[operator_index] if operator_index < len(parts) else ""


def _task_from_parts(parts: tuple[str, ...], root: Path | None, date_index: int | None) -> str:
    if date_index is None:
        return ""
    task_index = _task_index_from_parts(parts, date_index)
    if task_index is not None:
        return parts[task_index]
    if date_index > 0:
        return parts[date_index - 1]
    return root.name if root is not None else ""


def _task_index_from_parts(parts: tuple[str, ...], date_index: int | None) -> int | None:
    if date_index is None:
        return None
    if date_index >= 3 and _contains_known_robot_token(parts[date_index - 2]):
        return date_index - 3
    if date_index > 0:
        return date_index - 1
    return None


def _robot_from_layout_parts(parts: tuple[str, ...], date_index: int | None) -> str:
    if date_index is None or date_index < 3:
        return ""
    robot_type_folder = parts[date_index - 2]
    tokens = set(_name_tokens(robot_type_folder))
    for robot in _task_robot_tokens():
        if robot in tokens:
            return robot
    return ""


def _controller_from_episode_name(name: str) -> str:
    parts = name.split("_")
    if len(parts) >= 6 and parts[0] == "episode":
        return parts[-1]
    return ""


def _robot_from_episode_name(name: str) -> str:
    parts = name.split("_")
    if len(parts) >= 6 and parts[0] == "episode":
        return parts[-2]
    return ""


def _contains_known_robot_token(value: str) -> bool:
    return bool(set(_task_robot_tokens()) & set(_name_tokens(value)))


def _task_robot_tokens() -> list[str]:
    configured = config_value(["phase1_metadata", "task_robot_tokens"], {})
    if isinstance(configured, dict):
        return sorted(str(key).lower() for key in configured)
    return []


def _name_tokens(value: str) -> list[str]:
    tokens = []
    current = []
    for char in value.lower():
        if char.isalnum():
            current.append(char)
        elif current:
            tokens.append("".join(current))
            current = []
    if current:
        tokens.append("".join(current))
    return tokens


def _quality_report_rows(db_path: Path) -> list[dict]:
    init_db(db_path)
    states = load_all_states(db_path)
    findings_by_episode = _non_pass_findings_by_episode(db_path)
    rows = []
    for state in states:
        findings = findings_by_episode.get(str(state.episode_path), [])
        failed_checks = sorted({item[0] for item in findings})
        rows.append(
            {
                "episode_path": str(state.episode_path),
                "task": state.task,
                "date": state.date,
                "operator": state.operator,
                "robot": state.robot,
                "controller": state.controller,
                "status": state.final_status,
                "severity": _max_severity_from_rows(findings),
                "reasons": ";".join(failed_checks),
                "checked_at": state.last_updated,
            }
        )
    return rows


def _excel_finding_rows(db_path: Path) -> list[dict]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT f.episode_path, f.phase, f.check_name, f.severity, f.status,
                   f.message, f.details,
                   e.task, e.date, e.operator, e.robot, e.controller
            FROM findings f
            LEFT JOIN episodes e ON e.episode_path = f.episode_path
            WHERE f.status != ?
            ORDER BY f.episode_path, f.phase, f.id
            """,
            ("pass",),
        ).fetchall()
    return [
        {
            "episode_path": row["episode_path"],
            "task": row["task"] or "",
            "date": row["date"] or "",
            "operator": row["operator"] or "",
            "robot": row["robot"] or "",
            "controller": row["controller"] or "",
            "phase": row["phase"],
            "check_name": row["check_name"],
            "severity": row["severity"],
            "status": row["status"],
            "message": row["message"],
            "details": _compact_json(row["details"]),
        }
        for row in rows
    ]


def _excel_summary_rows(episode_rows: list[dict], finding_rows: list[dict]) -> list[dict]:
    status_counts = {}
    for row in episode_rows:
        status = row.get("status") or "pending"
        status_counts[status] = status_counts.get(status, 0) + 1
    severity_counts = {}
    for row in finding_rows:
        severity = row.get("severity") or ""
        severity_counts[severity] = severity_counts.get(severity, 0) + 1
    rows = [
        {"metric": "episodes", "value": len(episode_rows)},
        {"metric": "non_pass_findings", "value": len(finding_rows)},
    ]
    for status in ("fail", "needs_review", "warning", "pass", "pending"):
        rows.append({"metric": f"episodes_{status}", "value": status_counts.get(status, 0)})
    for severity in ("critical", "major", "minor", "info"):
        rows.append({"metric": f"findings_{severity}", "value": severity_counts.get(severity, 0)})
    return rows


def _excel_task_rows(episode_rows: list[dict]) -> list[dict]:
    grouped: dict[str, dict[str, int]] = {}
    for row in episode_rows:
        task = row.get("task") or "(unknown)"
        status = row.get("status") or "pending"
        grouped.setdefault(task, {"fail": 0, "needs_review": 0, "warning": 0, "pass": 0, "pending": 0})
        grouped[task][status] = grouped[task].get(status, 0) + 1
    return [
        {
            "task": task,
            **counts,
            "total": sum(counts.values()),
        }
        for task, counts in sorted(grouped.items())
    ]


def _excel_issue_count_rows(finding_rows: list[dict]) -> list[dict]:
    counts: dict[str, int] = {}
    for row in finding_rows:
        check_name = row.get("check_name") or "(unknown)"
        counts[check_name] = counts.get(check_name, 0) + 1
    return [
        {"check_name": check_name, "count": count}
        for check_name, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]


def _write_sheet(worksheet: Any, rows: list[dict], fieldnames: list[str]) -> None:
    worksheet.append(fieldnames)
    for row in rows:
        worksheet.append([row.get(fieldname, "") for fieldname in fieldnames])


def _autosize_columns(worksheet: Any, get_column_letter: Any) -> None:
    for column_cells in worksheet.columns:
        max_len = 0
        column_index = column_cells[0].column
        for cell in column_cells[:2000]:
            max_len = max(max_len, len(str(cell.value or "")))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 10), 60)


def _compact_json(value: str | None) -> str:
    if not value:
        return ""
    try:
        return json.dumps(json.loads(value), ensure_ascii=False, separators=(",", ":"))
    except json.JSONDecodeError:
        return value


def _load_findings_for_episode(db_path: Path, episode_path: str) -> list[Finding]:
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT episode_path, phase, check_name, severity, status, message, details
            FROM findings
            WHERE episode_path = ?
            ORDER BY phase, id
            """,
            (episode_path,),
        ).fetchall()
    return [_finding_from_row(row) for row in rows]


def _non_pass_findings_by_episode(db_path: Path) -> dict[str, list[tuple[str, str, str]]]:
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT episode_path, check_name, severity, status
            FROM findings
            WHERE status != 'pass'
            """,
            (),
        ).fetchall()
    findings_by_episode: dict[str, list[tuple[str, str, str]]] = {}
    for episode_path, check_name, severity, status in rows:
        findings_by_episode.setdefault(episode_path, []).append((check_name, severity, status))
    return findings_by_episode


def _max_severity_from_rows(findings: list[tuple[str, str, str]], minimum: str = "info") -> str:
    highest = minimum
    for _, severity, _ in findings:
        if severity_rank(severity) > severity_rank(highest):
            highest = severity
    return highest


def _finding_from_row(row: sqlite3.Row) -> Finding:
    return Finding(
        episode_path=row["episode_path"],
        phase=row["phase"],
        check_name=row["check_name"],
        severity=row["severity"],
        status=row["status"],
        message=row["message"],
        details=_json_dict(row["details"]),
    )


def _finding_json(row: sqlite3.Row) -> dict:
    return {
        "episode_path": row["episode_path"],
        "phase": row["phase"],
        "check_name": row["check_name"],
        "severity": row["severity"],
        "status": row["status"],
        "message": row["message"],
        "details": _json_dict(row["details"]),
    }


def _summary_overview_lines(db_path: Path) -> list[str]:
    total = _scalar_query(db_path, "SELECT COUNT(*) FROM episodes WHERE ? = ?", ("all", "all"))
    lines = ["## Overview", "", f"Total episodes checked: {total}", ""]
    lines.extend(["| Status | Count |", "|---|---:|"])
    for status in STATUS_VALUES:
        count = _scalar_query(
            db_path,
            "SELECT COUNT(*) FROM episodes WHERE final_status = ?",
            (status,),
        )
        lines.append("| " + status + " | " + str(count) + " |")
    lines.append("")
    return lines


def _summary_by_task_lines(db_path: Path) -> list[str]:
    return _summary_group_lines(db_path, "task", "Status by Task")


def _summary_by_operator_lines(db_path: Path) -> list[str]:
    return _summary_group_lines(db_path, "operator", "Status by Operator")


def _summary_by_robot_lines(db_path: Path) -> list[str]:
    return _summary_group_lines(db_path, "robot", "Status by Robot")


def _summary_top_issues_lines(db_path: Path) -> list[str]:
    rows = _rows_query(
        db_path,
        """
        SELECT check_name, COUNT(DISTINCT episode_path) AS episode_count
        FROM findings
        WHERE status != ? AND ? = ?
        GROUP BY check_name
        ORDER BY episode_count DESC, check_name
        LIMIT 10
        """,
        ("pass", "all", "all"),
    )
    lines = ["## Top Issues", "", "| Check | Affected Episodes |", "|---|---:|"]
    lines.extend("| " + row["check_name"] + " | " + str(row["episode_count"]) + " |" for row in rows)
    if not rows:
        lines.append("| None | 0 |")
    lines.append("")
    return lines


def _summary_group_lines(db_path: Path, group_field: str, title: str) -> list[str]:
    rows = _group_status_rows(db_path, group_field)
    lines = ["## " + title, "", "| " + group_field.title() + " | Pass | Warning | Fail | Needs Review |"]
    lines.append("|---|---:|---:|---:|---:|")
    for group_value, counts in rows.items():
        lines.append(
            "| "
            + group_value
            + " | "
            + str(counts["pass"])
            + " | "
            + str(counts["warning"])
            + " | "
            + str(counts["fail"])
            + " | "
            + str(counts["needs_review"])
            + " |"
        )
    lines.append("")
    return lines


def _group_status_rows(db_path: Path, group_field: str) -> dict[str, dict[str, int]]:
    if group_field not in {"task", "operator", "robot"}:
        raise ValueError("unsupported group field")
    statements = {
        "task": "SELECT task AS group_value, final_status, COUNT(*) AS count FROM episodes WHERE ? = ? GROUP BY task, final_status ORDER BY task",
        "operator": "SELECT operator AS group_value, final_status, COUNT(*) AS count FROM episodes WHERE ? = ? GROUP BY operator, final_status ORDER BY operator",
        "robot": "SELECT robot AS group_value, final_status, COUNT(*) AS count FROM episodes WHERE ? = ? GROUP BY robot, final_status ORDER BY robot",
    }
    rows = _rows_query(db_path, statements[group_field], ("all", "all"))
    grouped: dict[str, dict[str, int]] = {}
    for row in rows:
        key = row["group_value"] or ""
        grouped.setdefault(key, {status: 0 for status in STATUS_VALUES})
        if row["final_status"] in STATUS_VALUES:
            grouped[key][row["final_status"]] = row["count"]
    return grouped


def _summary_problem_patterns_lines(db_path: Path) -> list[str]:
    lines = [
        "## Problem Patterns",
        "",
        "Groups with 30%+ flagged rate (fail or needs_review), by task.",
        "",
    ]
    tasks = _summary_tasks(db_path)
    pattern_rows = _problem_pattern_rows(db_path)
    by_task: dict[str, list[sqlite3.Row]] = {}
    for row in pattern_rows:
        by_task.setdefault(row["task"] or "", []).append(row)
    for task in tasks:
        lines.append("### " + task)
        lines.append("")
        rows = by_task.get(task, [])
        if not rows:
            lines.append("No groups with 30%+ flagged rate for this task.")
            lines.append("")
            continue
        lines.extend(_problem_pattern_table_lines(db_path, rows))
    if not tasks:
        lines.append("No episodes found.")
        lines.append("")
    return lines


def _summary_appendix_lines(db_path: Path) -> list[str]:
    lines = ["## Appendix: Flagged Episodes", ""]
    tasks = _flagged_tasks(db_path)
    if not tasks:
        lines.append("No fail or needs_review episodes.")
        lines.append("")
        return lines
    findings_by_episode = _appendix_findings_by_episode(db_path)
    for task in tasks:
        lines.extend(_appendix_task_lines(db_path, task, findings_by_episode))
    return lines


def _summary_tasks(db_path: Path) -> list[str]:
    rows = _rows_query(
        db_path,
        """
        SELECT DISTINCT task
        FROM episodes
        WHERE ? = ?
        ORDER BY task
        """,
        ("all", "all"),
    )
    return [row["task"] or "" for row in rows]


def _problem_pattern_rows(db_path: Path) -> list[sqlite3.Row]:
    return _rows_query(
        db_path,
        """
        SELECT e.task, e.date, e.operator,
               COUNT(*) AS total,
               SUM(CASE WHEN e.final_status = ? THEN 1 ELSE 0 END) AS fail_count,
               SUM(CASE WHEN e.final_status = ? THEN 1 ELSE 0 END) AS review_count
        FROM episodes e
        WHERE ? = ?
        GROUP BY e.task, e.date, e.operator
        HAVING (fail_count + review_count) * 1.0 / total >= ?
        ORDER BY e.task, (fail_count + review_count) * 1.0 / total DESC
        """,
        ("fail", "needs_review", "all", "all", 0.30),
    )


def _problem_pattern_table_lines(db_path: Path, rows: list[sqlite3.Row]) -> list[str]:
    lines = ["| Date | Operator | Total | Fail | Review | Rate | Top Issues |"]
    lines.append("|---|---|---:|---:|---:|---:|---|")
    for row in rows:
        total = row["total"] or 0
        fail_count = row["fail_count"] or 0
        review_count = row["review_count"] or 0
        rate = (fail_count + review_count) / total if total else 0.0
        top_issues = ", ".join(_top_issues_for_group(db_path, row["task"], row["date"], row["operator"]))
        lines.append(
            "| "
            + str(row["date"] or "")
            + " | "
            + str(row["operator"] or "")
            + " | "
            + str(total)
            + " | "
            + str(fail_count)
            + " | "
            + str(review_count)
            + " | "
            + f"{rate:.0%}"
            + " | "
            + top_issues
            + " |"
        )
    lines.append("")
    return lines


def _top_issues_for_group(db_path: Path, task: str, date: str, operator: str) -> list[str]:
    rows = _rows_query(
        db_path,
        """
        SELECT f.check_name, COUNT(DISTINCT f.episode_path) AS episode_count
        FROM findings f
        JOIN episodes e ON e.episode_path = f.episode_path
        WHERE e.task = ?
          AND e.date = ?
          AND e.operator = ?
          AND e.final_status IN (?, ?)
          AND f.status != ?
        GROUP BY f.check_name
        ORDER BY episode_count DESC, f.check_name
        LIMIT 3
        """,
        (task, date, operator, "fail", "needs_review", "pass"),
    )
    return [row["check_name"] for row in rows]


def _flagged_tasks(db_path: Path) -> list[str]:
    rows = _rows_query(
        db_path,
        """
        SELECT DISTINCT task
        FROM episodes
        WHERE final_status IN (?, ?)
        ORDER BY task
        """,
        ("fail", "needs_review"),
    )
    return [row["task"] or "" for row in rows]


def _appendix_task_lines(
    db_path: Path, task: str, findings_by_episode: dict[str, list[sqlite3.Row]]
) -> list[str]:
    lines = ["### Task: " + task, ""]
    lines.extend(_appendix_status_examples(db_path, task, "fail", "Fail examples", findings_by_episode))
    lines.extend(_appendix_status_examples(db_path, task, "needs_review", "Needs review examples", findings_by_episode))
    fail_count = _task_status_count(db_path, task, "fail")
    review_count = _task_status_count(db_path, task, "needs_review")
    lines.append(
        f"*(Task total: {fail_count} fail, {review_count} needs_review. Full list in quality_findings.jsonl)*"
    )
    lines.append("")
    return lines


def _appendix_status_examples(
    db_path: Path,
    task: str,
    status: str,
    title: str,
    findings_by_episode: dict[str, list[sqlite3.Row]],
) -> list[str]:
    lines = ["#### " + title, ""]
    rows = _appendix_episode_rows(db_path, task, status)
    if not rows:
        lines.append("None.")
        lines.append("")
        return lines
    for row in rows:
        lines.extend(
            _appendix_episode_detail_lines(
                row["episode_path"], row["final_status"], findings_by_episode
            )
        )
    return lines


def _appendix_episode_rows(db_path: Path, task: str, status: str) -> list[sqlite3.Row]:
    return _rows_query(
        db_path,
        """
        SELECT episode_path, final_status
        FROM episodes
        WHERE task = ? AND final_status = ?
        ORDER BY episode_path
        LIMIT 5
        """,
        (task, status),
    )


def _appendix_episode_detail_lines(
    episode_path: str, final_status: str, findings_by_episode: dict[str, list[sqlite3.Row]]
) -> list[str]:
    name = Path(episode_path).name
    lines = [f"**[{final_status.upper()}] {name}**"]
    for line in _appendix_finding_lines(findings_by_episode.get(episode_path, [])):
        lines.append(line)
    lines.append("")
    return lines


def _appendix_findings_by_episode(db_path: Path) -> dict[str, list[sqlite3.Row]]:
    rows = _rows_query(
        db_path,
        """
        SELECT episode_path, phase, check_name, severity, details
        FROM findings
        WHERE status != ?
        ORDER BY episode_path, phase,
        CASE severity
        WHEN 'critical' THEN 0
        WHEN 'major' THEN 1
        WHEN 'minor' THEN 2
        ELSE 3 END
        """,
        ("pass",),
    )
    findings_by_episode: dict[str, list[sqlite3.Row]] = {}
    for row in rows:
        findings_by_episode.setdefault(row["episode_path"], []).append(row)
    return findings_by_episode


def _appendix_finding_lines(rows: list[sqlite3.Row]) -> list[str]:
    lines = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        details = _json_dict(row["details"]) if isinstance(row["details"], str) else {}
        modality = details.get("modality", "")
        key = (row["check_name"], modality)
        if key in seen:
            continue
        seen.add(key)
        extra = _format_finding_extra_md(details)
        modality_part = f" `{modality}`" if modality else ""
        extra_part = f" — {extra}" if extra else ""
        lines.append(f"- p{row['phase']} `{row['severity']}` **{row['check_name']}**{modality_part}{extra_part}")
    return lines


def _task_status_count(db_path: Path, task: str, status: str) -> int:
    return int(
        _scalar_query(
            db_path,
            "SELECT COUNT(*) FROM episodes WHERE task = ? AND final_status = ?",
            (task, status),
        )
    )


def _format_finding_extra_md(details: dict) -> str:
    """Format a short extra info string from finding details for markdown output."""
    if "drop_ratio" in details:
        return f"drop_ratio={details['drop_ratio']:.1%}"
    if "max_consecutive_drops" in details:
        return f"max_consecutive={details['max_consecutive_drops']}"
    if "violation_ratio" in details:
        return f"ratio={details['violation_ratio']:.1%}"
    if "duplicate_ratio" in details:
        return f"ratio={details['duplicate_ratio']:.1%}"
    if "deviation_ratio" in details:
        return f"deviation={details['deviation_ratio']:.1%}"
    if "spread_ms" in details:
        return f"spread={details['spread_ms']:.0f}ms"
    if "iqr_distance" in details:
        return f"iqr_distance={details['iqr_distance']:.1f}"
    if "threshold" in details:
        return f"threshold={details['threshold']}"
    return ""


def _scalar_query(db_path: Path, statement: str, params: tuple) -> Any:
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(statement, params).fetchone()
    return row[0] if row else 0


def _rows_query(db_path: Path, statement: str, params: tuple) -> list[sqlite3.Row]:
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        return conn.execute(statement, params).fetchall()
